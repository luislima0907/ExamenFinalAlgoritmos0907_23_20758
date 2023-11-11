import openpyxl
import sys
import re

libro = openpyxl.load_workbook("vehiculos.xlsx")

hoja = libro['listado']

hoja['A1'].value = "Id"
hoja['B1'].value = "Codigo"
hoja['C1'].value = "Marca"
hoja['D1'].value = "Modelo"
hoja['E1'].value = "Precio"
hoja['F1'].value = "kilometraje"

libro.save("vehiculos.xlsx")

if sys.argv[1] == "ayuda":
    print('''1. Para ingresar vehiculos debemos poner el comando: python mega_autos.py "Ingresar" "J89,Honda,carro,87000.00,20"\n 2. Para modificar vehiculos tenemos que buscar por ID y colocamos la nueva informacion: py mega_auto.py "vehiculos.xlsx" "Modificar" "D59,Suzuki,moto,97000.60,40" 2 \n 3. Para Listar los vehiculos debemos de escribir lo siguiente: py mega_auto.py "Listar" \n 4. Para eliminar los vehiculos debemos de buscar por su ID y escribir el vehiculo completo: py mega_auto.py "vehiculos.xlsx" "Eliminar" "D59,Suzuki,moto,97000.60,40" 2 \n ''')

if sys.argv[1] == "Ingresar":
    
    #Abrimos el libro
    libro = openpyxl.load_workbook("vehiculos.xlsx")

    #abrimos una hoja
    hoja = libro['listado']
    
    vehiculo_ingresado = sys.argv[2]

    def el_vehiculo_es_valido_con_expresiones(vehiculo_ingresado):
        
        #El metodo re.compile(r'') nos sirve para crear el formato que deseemos en el programa.
        expresion_para_ingresar_vehiculos = re.compile(r'\w+\,\w+\,\w+\,\d+.\d\d\,\d+')
        
        # buscamos las coincidencias con el fomato introducido
        coincidencias = expresion_para_ingresar_vehiculos.search(vehiculo_ingresado)
        
        #Si no hay coincidencias entonces el programa no continua
        if coincidencias == None:
            return False
        
        # el metodo .group() devuelve la primera coincidencia encontrada
        grupo = coincidencias.group()
        return grupo == vehiculo_ingresado

    #Nos devuelve True o False
    print(el_vehiculo_es_valido_con_expresiones(vehiculo_ingresado))
    
    informacion_separada = vehiculo_ingresado.split(',')
    print(informacion_separada)
    codigo = informacion_separada[0]
    marca = informacion_separada[1]
    modelo = informacion_separada[2]
    precio = informacion_separada[3]
    kilometraje = int(informacion_separada[4])
    
    proxima_fila = hoja.max_row + 1    

    datos_reales = [
        {
            "Id": proxima_fila,
            "Codigo": codigo,
            "Marca": marca,
            "Modelo": modelo,
            "Precio": precio,
            "Kilometraje": kilometraje,
        }
    ]

    proxima_fila = hoja.max_row + 1    

    for vehiculos in datos_reales:
        hoja[f'A{proxima_fila}'].value = vehiculos["Id"]
        hoja[f'B{proxima_fila}'].value = vehiculos["Codigo"]
        hoja[f'C{proxima_fila}'].value = vehiculos["Marca"]
        hoja[f'D{proxima_fila}'].value = vehiculos["Modelo"]
        hoja[f'E{proxima_fila}'].value = vehiculos["Precio"]
        hoja[f'F{proxima_fila}'].value = vehiculos["Kilometraje"]

if sys.argv[1] == "vehiculos.xlsx" and sys.argv[2] == "Modificar":
        #Abrimos el libro
    libro = openpyxl.load_workbook("vehiculos.xlsx")

    #abrimos una hoja
    hoja = libro['listado']
    
    nuevo_vehiculo_ingresado = sys.argv[3]
    
    def el_vehiculo_es_valido_con_expresiones_nuevo(nuevo_vehiculo_ingresado):
        
        #El metodo re.compile(r'') nos sirve para crear el formato que deseemos en el programa.
        expresion_para_ingresar_vehiculos_nuevos = re.compile(r'\w+\,\w+\,\w+\,\d+.\d\d\,\d+')
        
        # buscamos las coincidencias con el fomato introducido
        coincidencias = expresion_para_ingresar_vehiculos_nuevos.search(nuevo_vehiculo_ingresado)
        
        #Si no hay coincidencias entonces el programa no continua
        if coincidencias == None:
            return False
        
        # el metodo .group() devuelve la primera coincidencia encontrada
        grupo = coincidencias.group()
        return grupo == nuevo_vehiculo_ingresado
    
    nueva_informacion_separada = nuevo_vehiculo_ingresado.split(',')
    print(nueva_informacion_separada)
    nuevo_codigo = nueva_informacion_separada[0]
    nueva_marca = nueva_informacion_separada[1]
    nuevo_modelo = nueva_informacion_separada[2]
    nuevo_precio = nueva_informacion_separada[3]
    nuevo_kilometraje = nueva_informacion_separada[4]

    #Fila donde encontramos el ID
    fila_encontrada = 0
    for numero_fila in range(2, hoja.max_row + 1):
        if hoja[f'A{numero_fila}'].value == int(sys.argv[4]):
            fila_encontrada = numero_fila
            break

    if fila_encontrada > 0:
        hoja[f'B{fila_encontrada}'].value = nuevo_codigo
        hoja[f'C{fila_encontrada}'].value = nueva_marca
        hoja[f'D{fila_encontrada}'].value = nuevo_modelo
        hoja[f'E{fila_encontrada}'].value = nuevo_precio
        hoja[f'F{fila_encontrada}'].value = nuevo_kilometraje
                
        datos_reales = [
            {
                "Codigo": nuevo_codigo,
                "Marca": nueva_marca,
                "Modelo": nuevo_modelo,
                "Precio": nuevo_precio,
                "Kilometraje": nuevo_kilometraje,
            }
        ]

        for vehiculos in datos_reales:
            hoja[f'B{fila_encontrada}'].value = vehiculos["Codigo"]
            hoja[f'C{fila_encontrada}'].value = vehiculos["Marca"]
            hoja[f'D{fila_encontrada}'].value = vehiculos["Modelo"]
            hoja[f'E{fila_encontrada}'].value = vehiculos["Precio"]
            hoja[f'F{fila_encontrada}'].value = vehiculos["Kilometraje"]
        
    else:
        print("Vehiculo no encontrado")

libro.save(sys.argv[1])

if sys.argv[1] == "vehiculos.xlsx" and sys.argv[2] == "Listar":
    #Abrimos el libro
    libro = openpyxl.load_workbook("vehiculos.xlsx")

    #abrimos una hoja
    hoja = libro['listado']
    
    def leer_vehiculos():
        veiculos = []
        for numero_fila in range(2, hoja.max_row + 1):
            veiculos.append({
                "Codigo": hoja[f"B{numero_fila}"].value,
                "Marca": hoja[f"C{numero_fila}"].value,
                "Modelo":  hoja[f"D{numero_fila}"].value,
                "Precio": hoja[f"E{numero_fila}"].value,
                "Kilometraje": hoja[f"F{numero_fila}"].value
            })
        print(veiculos)
    leer_vehiculos()   

if sys.argv[1] == "vehiculos.xlsx" and sys.argv[2] == "Eliminar":
    #Abrimos el libro
    libro = openpyxl.load_workbook("vehiculos.xlsx")

    #abrimos una hoja
    hoja = libro['listado']
    
    eliminar_vehiculo_ingresado = sys.argv[3]
    
    def eliminar_vehiculo__con_expresiones_nuevo(eliminar_vehiculo_ingresado):
        
        #El metodo re.compile(r'') nos sirve para crear el formato que deseemos en el programa.
        expresion_para_eliminar_vehiculos = re.compile(r'\w+\,\w+\,\w+\,\d+.\d\d\,\d+')
        
        # buscamos las coincidencias con el fomato introducido
        coincidencias = expresion_para_eliminar_vehiculos.search(eliminar_vehiculo_ingresado)
        
        #Si no hay coincidencias entonces el programa no continua
        if coincidencias == None:
            return False
        
        # el metodo .group() devuelve la primera coincidencia encontrada
        grupo = coincidencias.group()
        return grupo == eliminar_vehiculo_ingresado
    
    nueva_informacion_separada = eliminar_vehiculo_ingresado.split(',')
    print(nueva_informacion_separada)
    eliminar_codigo = nueva_informacion_separada[0]
    eliminar_marca = nueva_informacion_separada[1]
    eliminar_modelo = nueva_informacion_separada[2]
    eliminar_precio = nueva_informacion_separada[3]
    eliminar_kilometraje = nueva_informacion_separada[4]
    
    fila_encontrada = 0
    for numero_fila in range(2, hoja.max_row + 1):
        if hoja[f'A{numero_fila}'].value == int(sys.argv[4]):
            fila_encontrada = numero_fila
            break
    
    if fila_encontrada > 0:
        hoja[f'B{fila_encontrada}'].value = ""
        hoja[f'C{fila_encontrada}'].value = ""
        hoja[f'D{fila_encontrada}'].value = ""
        hoja[f'E{fila_encontrada}'].value = ""
        hoja[f'F{fila_encontrada}'].value = ""
                        
        datos_reales = [
            {
                "Id": "",
                "Codigo": "",
                "Marca": "",
                "Modelo": "",
                "Precio": "",
                "Kilometraje": "",
            }
        ]
    libro.save(sys.argv[1])

libro.save("vehiculos.xlsx")